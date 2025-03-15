from PyPDF2 import PdfWriter, PdfReader
import io
import os
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.pdfmetrics import stringWidth
import openpyxl 
from reportlab.pdfbase.ttfonts import TTFont  

pdfmetrics.registerFont(TTFont('Verdana', 'Verdana.ttf'))
print ("Please put the list snd the blank pdf file on assets folder")
foldr = input('Please Enter the output folder name: ')
path  = os.path.join('./', foldr)
print (path)
if not os.path.exists(path):
  os.mkdir(path)
  print("Folder %s created!" % path)
else:
  print("Folder %s already exists" % path)
fn = input('Please Enter the list filename: ')
#wb = openpyxl.load_workbook(fn)
wb = openpyxl.load_workbook(os.path.join('./assets',fn+'.xlsx'))
ws = wb.active
filepdf = input('Please Enter the PDF filename: ')
fpdf = (os.path.join('./assets',filepdf+'.pdf'))
hdrs = input('List has headers (Yes/No): ')
if hdrs == 'Yes':
    s_row = 2
else:   
    s_row = 1
n_rows= ws.max_row


for i in range(s_row, n_rows):

    packet = io.BytesIO()
    text = ws.cell(row=i,column=1).value + " " +ws.cell(row=i,column=2).value
    can = canvas.Canvas(packet)
    can.setPageSize((210 * mm, 297 * mm ))
    can.setFont('Verdana', 28)
    x = can._pagesize[0] / 2
    can.drawCentredString(x, 440, text)
    can.save()
    #move to the beginning of the StringIO buffer
    packet.seek(0)
    # create a new PDF with Reportlab
    new_pdf = PdfReader(packet)
    # read your existing PDF
    existing_pdf = PdfReader(open(fpdf, "rb"))
    output = PdfWriter()
    # add the "watermark" (which is the new pdf) on the existing page
    page = existing_pdf.pages[0]
    page.merge_page(new_pdf.pages[0])
    output.add_page(page)
    # finally, write "output" to a real file
    flnm = text+'.pdf'
    output_stream = open(os.path.join(path,flnm), 'wb')
    output.write(output_stream)
    output_stream.close()
